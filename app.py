from flask import Flask, render_template, request, send_file, jsonify
import pandas as pd
import numpy as np
import os
import io
import re
from datetime import datetime
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import threading
import time

app = Flask(__name__)

# Global variable to track processing state
processing_status = {
    "is_processing": False,
    "start_time": None,
    "complete_time": None,
    "is_complete": False,
    "filename": None,
    "sheets_processed": [],
    "error": None
}

def log_message(message):
    """
    Log a message with timestamp to console and log file
    """
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
    log_entry = f"[{timestamp}] {message}"
    print(log_entry)
    
    # Also write to log file
    log_dir = os.path.join(os.getcwd(), 'logs')
    os.makedirs(log_dir, exist_ok=True)
    
    log_file = os.path.join(log_dir, f"app_{datetime.now().strftime('%Y-%m-%d')}.log")
    with open(log_file, 'a', encoding='utf-8') as f:
        f.write(log_entry + "\n")
    
    return log_entry

def clean_excel_data(df, data_type='delay'):
    """
    Clean Excel data based on the data type (delay or cycle)
    """
    if data_type == 'delay':
        return clean_delay_data(df)
    elif data_type == 'cycle':
        return clean_cycle_data(df)
    else:
        raise ValueError("Invalid data type. Must be 'delay' or 'cycle'")

def clean_delay_data(df):
    """
    Clean delay data from the Excel file
    """
    # Create a new DataFrame with desired structure
    cleaned_df = pd.DataFrame(columns=[
        'Unit', 'Start', 'Finish', 'Dur', 'Desc', 'Delay Type', 'Category'
    ])
    
    # Map columns from original data to new format
    if 'Machine' in df.columns:
        cleaned_df['Unit'] = df['Machine']
    elif 'Unit' in df.columns:
        cleaned_df['Unit'] = df['Unit']
    
    # Handle date/time columns
    if 'Start Date' in df.columns and 'Finish Date' in df.columns:
        # Helper function to parse WIT date format to datetime object
        def parse_wit_date(date_str):
            if pd.isna(date_str):
                return None
                
            # Extract date components from strings like "Fri Mar 28 07:25:25 WIT 2025"
            pattern = r'(\w{3}) (\w{3}) (\d{1,2}) (\d{2}:\d{2}:\d{2}) WIT (\d{4})'
            match = re.match(pattern, str(date_str))
            
            if match:
                _, month, day, time, year = match.groups()
                
                # Convert month name to number
                month_map = {
                    'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                    'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
                }
                month_num = month_map.get(month, 1)
                
                # Format date string for parsing
                date_str = f"{year}-{month_num:02d}-{int(day):02d} {time}"
                
                # Return as datetime object
                try:
                    return pd.to_datetime(date_str)
                except:
                    return None
            
            # If pattern doesn't match, try direct pandas conversion
            try:
                return pd.to_datetime(date_str)
            except:
                return None
        
        # Apply the parsing function
        cleaned_df['Start'] = df['Start Date'].apply(parse_wit_date)
        cleaned_df['Finish'] = df['Finish Date'].apply(parse_wit_date)
        
        # Calculate duration as difference between finish and start times (in hours)
        durations = []
        for i in range(len(cleaned_df)):
            start = cleaned_df['Start'].iloc[i]
            finish = cleaned_df['Finish'].iloc[i]
            
            if pd.notna(start) and pd.notna(finish):
                try:
                    # Calculate duration in hours
                    duration_seconds = (finish - start).total_seconds()
                    duration_hours = duration_seconds / 3600
                    durations.append(round(duration_hours, 2))
                except:
                    # If calculation fails, use original Duration if available
                    if 'Duration' in df.columns and i < len(df['Duration']):
                        try:
                            # Handle format like "9,545 s"
                            dur_str = df['Duration'].iloc[i]
                            match = re.match(r'([\d,]+\.?\d*)\s*s', str(dur_str))
                            if match:
                                # Remove commas and convert to float
                                seconds = float(match.group(1).replace(',', ''))
                                # Convert to hours
                                durations.append(round(seconds / 3600, 2))
                            else:
                                # Try direct numeric conversion if no unit
                                durations.append(round(float(str(dur_str).replace(',', '')) / 3600, 2))
                        except:
                            durations.append(None)
                    else:
                        durations.append(None)
            else:
                # If either start or finish is missing, use original Duration if available
                if 'Duration' in df.columns and i < len(df['Duration']):
                    try:
                        # Handle format like "9,545 s"
                        dur_str = df['Duration'].iloc[i]
                        match = re.match(r'([\d,]+\.?\d*)\s*s', str(dur_str))
                        if match:
                            # Remove commas and convert to float
                            seconds = float(match.group(1).replace(',', ''))
                            # Convert to hours
                            durations.append(round(seconds / 3600, 2))
                        else:
                            # Try direct numeric conversion if no unit
                            durations.append(round(float(str(dur_str).replace(',', '')) / 3600, 2))
                    except:
                        durations.append(None)
                else:
                    durations.append(None)
                    
        cleaned_df['Dur'] = durations
    else:
        # Process duration (convert from seconds to hours with 2 decimal places)
        if 'Duration' in df.columns:
            def convert_duration(dur_str):
                if pd.isna(dur_str):
                    return None
                    
                # Handle format like "9,545 s"
                match = re.match(r'([\d,]+\.?\d*)\s*s', str(dur_str))
                if match:
                    # Remove commas and convert to float
                    seconds = float(match.group(1).replace(',', ''))
                    # Convert to hours
                    return round(seconds / 3600, 2)
                
                # Try direct numeric conversion if no unit
                try:
                    return round(float(str(dur_str).replace(',', '')) / 3600, 2)
                except:
                    return None
            
            cleaned_df['Dur'] = df['Duration'].apply(convert_duration)
    
    # Copy description
    if 'Description' in df.columns:
        cleaned_df['Desc'] = df['Description']
    elif 'Desc' in df.columns:
        cleaned_df['Desc'] = df['Desc']
    
    # Copy delay type
    if 'Delay Type' in df.columns:
        cleaned_df['Delay Type'] = df['Delay Type']
    
    # Copy or determine category 
    if 'Category' in df.columns:
        # Direct mapping to standardize category names
        def standardize_category(category):
            if pd.isna(category):
                return None
                
            category = str(category).strip().upper()
            
            # Map to the exact format from the dropdown in the image
            if category in ['DELAY']:
                return 'DELAY'
            elif category in ['EXTENDED LOSS', 'EXTENDED']:
                return 'EXTENDED LOSS'
            elif category in ['PLANNED DOWN']:
                return 'PLANNED DOWN'
            elif category in ['STANDBY']:
                return 'STANDBY'
            elif category in ['UNPLANNED DOWN']:
                return 'UNPLANNED DOWN'
            
            return category
            
        cleaned_df['Category'] = df['Category'].apply(standardize_category)
    # Determine category based on delay type if Category column doesn't exist
    elif 'Delay Type' in df.columns:
        def get_category(delay_type):
            if pd.isna(delay_type):
                return None
                
            delay_type = str(delay_type).strip().upper()
            
            # Check for exact matches first
            if delay_type == 'DELAY':
                return 'DELAY'
            elif delay_type == 'EXTENDED LOSS':
                return 'EXTENDED LOSS'
            elif delay_type == 'PLANNED DOWN':
                return 'PLANNED DOWN'
            elif delay_type == 'STANDBY':
                return 'STANDBY'
            elif delay_type == 'UNPLANNED DOWN':
                return 'UNPLANNED DOWN'
            
            # Then check for prefix patterns, handling both with and without spaces
            if delay_type.startswith('D-') or delay_type.startswith('D -'):
                return 'DELAY'
            elif delay_type.startswith('S-') or delay_type.startswith('S -'):
                return 'STANDBY'
            elif delay_type.startswith('UX-') or delay_type.startswith('UX -'):
                return 'UNPLANNED DOWN'
            elif delay_type.startswith('X-') or delay_type.startswith('X -'):
                return 'PLANNED DOWN'
            elif delay_type.startswith('XX-') or delay_type.startswith('XX -'):
                return 'EXTENDED LOSS'
            return None
        
        cleaned_df['Category'] = df['Delay Type'].apply(get_category)
    
    return cleaned_df

def clean_cycle_data(df):
    """
    Clean cycle data from Excel file
    """
    # Create a new DataFrame that preserves original columns
    cleaned_df = pd.DataFrame()
    
    # Keep original Unit column
    if 'Unit' in df.columns:
        cleaned_df['Unit'] = df['Unit']
    
    # Keep original Operator column
    if 'Operator' in df.columns:
        cleaned_df['Operator'] = df['Operator']
    
    # Parse and store start time for duration calculation
    start_times = []
    
    # Handle Start time column, formatting to match delay data format
    if 'Start time' in df.columns:
        def format_date(date_str):
            if pd.isna(date_str):
                return None
            
            try:
                # Extract date components from strings like "Thu Mar 27 22:12:49 WIT 2025"
                pattern = r'(\w{3}) (\w{3}) (\d{1,2}) (\d{2}:\d{2}:\d{2}) WIT (\d{4})'
                match = re.match(pattern, str(date_str))
                
                if match:
                    _, month, day, time, year = match.groups()
                    
                    # Convert month name to number
                    month_map = {
                        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
                    }
                    month_num = month_map.get(month, 1)
                    
                    # Format date string for parsing
                    date_str = f"{year}-{month_num:02d}-{int(day):02d} {time}"
                    
                    # Return as datetime object
                    dt_obj = pd.to_datetime(date_str)
                    start_times.append(dt_obj)
                    return dt_obj
                
                # If pattern doesn't match, try using pandas
                date = pd.to_datetime(date_str)
                start_times.append(date)
                return date
            except:
                start_times.append(None)
                return None
            
        cleaned_df['Start'] = df['Start time'].apply(format_date)
    elif 'Start' in df.columns:  # Fallback for different column name
        def format_start_date(x):
            if pd.notna(x):
                try:
                    date = pd.to_datetime(x)
                    start_times.append(date)
                    return date
                except:
                    start_times.append(None)
                    return None
            else:
                start_times.append(None)
                return None
                
        cleaned_df['Start'] = df['Start'].apply(format_start_date)
    
    # Parse and store finish time for duration calculation
    finish_times = []
    
    # Handle Finish Time column
    if 'Finish Time' in df.columns:
        def format_date(date_str):
            if pd.isna(date_str):
                return None
            
            try:
                # Extract date components from strings like "Fri Mar 28 03:06:06 WIT 2025"
                pattern = r'(\w{3}) (\w{3}) (\d{1,2}) (\d{2}:\d{2}:\d{2}) WIT (\d{4})'
                match = re.match(pattern, str(date_str))
                
                if match:
                    _, month, day, time, year = match.groups()
                    
                    # Convert month name to number
                    month_map = {
                        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
                        'Jul': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
                    }
                    month_num = month_map.get(month, 1)
                    
                    # Format date string for parsing
                    date_str = f"{year}-{month_num:02d}-{int(day):02d} {time}"
                    
                    # Return as datetime object
                    dt_obj = pd.to_datetime(date_str)
                    finish_times.append(dt_obj)
                    return dt_obj
                
                # If pattern doesn't match, try using pandas
                date = pd.to_datetime(date_str)
                finish_times.append(date)
                return date
            except:
                finish_times.append(None)
                return None
            
        cleaned_df['Finish'] = df['Finish Time'].apply(format_date)
    elif 'Finish' in df.columns:  # Fallback for different column name
        def format_finish_date(x):
            if pd.notna(x):
                try:
                    date = pd.to_datetime(x)
                    finish_times.append(date)
                    return date
                except:
                    finish_times.append(None)
                    return None
            else:
                finish_times.append(None)
                return None
                
        cleaned_df['Finish'] = df['Finish'].apply(format_finish_date)
    
    # Calculate duration as difference between finish and start times (in hours)
    if len(start_times) == len(finish_times) and len(start_times) > 0:
        durations = []
        for i in range(len(start_times)):
            start = start_times[i]
            finish = finish_times[i]
            
            if pd.notna(start) and pd.notna(finish):
                try:
                    # Calculate duration in hours
                    duration_seconds = (finish - start).total_seconds()
                    duration_hours = duration_seconds / 3600
                    durations.append(round(duration_hours, 2))
                except:
                    # If calculation fails, use original Dur if available
                    if 'Dur' in df.columns and i < len(df['Dur']):
                        try:
                            durations.append(round(float(str(df['Dur'].iloc[i]).replace(',', '.')), 2))
                        except:
                            durations.append(df['Dur'].iloc[i] if i < len(df['Dur']) else None)
                    else:
                        durations.append(None)
            else:
                # If either start or finish is missing, use original Dur if available
                if 'Dur' in df.columns and i < len(df['Dur']):
                    try:
                        durations.append(round(float(str(df['Dur'].iloc[i]).replace(',', '.')), 2))
                    except:
                        durations.append(df['Dur'].iloc[i] if i < len(df['Dur']) else None)
                else:
                    durations.append(None)
        
        cleaned_df['Dur'] = durations
    else:
        # Process duration (ensure 2 decimal places)
        if 'Dur' in df.columns:
            def format_duration(dur_val):
                if pd.isna(dur_val):
                    return None
                    
                try:
                    # Convert to float and round to 2 decimal places
                    return round(float(str(dur_val).replace(',', '.')), 2)
                except:
                    return dur_val
            
            cleaned_df['Dur'] = df['Dur'].apply(format_duration)
    
    # Keep original Source column
    if 'Source' in df.columns:
        cleaned_df['Source'] = df['Source']
    
    # Keep original Destination column
    if 'Destination' in df.columns:
        cleaned_df['Destination'] = df['Destination']
    
    # Handle description column if present
    if 'Desc' in df.columns:
        cleaned_df['Desc'] = df['Desc']
    
    # Handle Delay Type and Category if present
    if 'Delay Type' in df.columns:
        cleaned_df['Delay Type'] = df['Delay Type']
    
    if 'Category' in df.columns:
        cleaned_df['Category'] = df['Category']
    
    return cleaned_df

def create_performance_sheet(delay_df, cycle_df):
    """
    Create Performance sheet based on data from Delay and Cycle sheets
    """
    if delay_df is None or cycle_df is None:
        return None
    
    # Get unique units from both sheets
    units = set()
    if 'Unit' in cycle_df.columns:
        units.update(cycle_df['Unit'].dropna().unique())
    if 'Unit' in delay_df.columns:
        units.update(delay_df['Unit'].dropna().unique())
    
    units = sorted(list(units))
    
    # Create DataFrame for Performance sheet with exact column names as in the image
    performance_df = pd.DataFrame(columns=[
        'Unit', 'Operating hrs', 'DELAY', 'EXTENDED LOSS', 'PLANNED DOWN', 
        'STANDBY', 'UNPLANNED DOWN', 'Grand Total', 'PA', 'UA'
    ])
    
    # Add units to the DataFrame
    performance_df['Unit'] = units
    
    # Calculate Operating hrs for each unit from Cycle data
    operating_hrs = {}
    for unit in units:
        if 'Unit' in cycle_df.columns and 'Dur' in cycle_df.columns:
            unit_data = cycle_df[cycle_df['Unit'] == unit]
            total_dur = unit_data['Dur'].sum()
            operating_hrs[unit] = round(total_dur, 2) if not pd.isna(total_dur) else 0.0
        else:
            operating_hrs[unit] = 0.0
    
    performance_df['Operating hrs'] = performance_df['Unit'].map(operating_hrs).fillna(0.0)
    
    # Calculate delay categories for each unit using exact category names
    categories = ['DELAY', 'EXTENDED LOSS', 'PLANNED DOWN', 'STANDBY', 'UNPLANNED DOWN']
    
    for category in categories:
        category_values = {}
        for unit in units:
            if 'Unit' in delay_df.columns and 'Category' in delay_df.columns and 'Dur' in delay_df.columns:
                # Filter delay data for this unit and category
                unit_category_data = delay_df[(delay_df['Unit'] == unit) & (delay_df['Category'] == category)]
                total_dur = unit_category_data['Dur'].sum()
                category_values[unit] = round(total_dur, 2) if not pd.isna(total_dur) else 0.0
            else:
                category_values[unit] = 0.0
        
        performance_df[category] = performance_df['Unit'].map(category_values).fillna(0.0)
    
    # Calculate Grand Total (sum of Operating hrs and all delay categories)
    performance_df['Grand Total'] = (performance_df['Operating hrs'] + 
                                    performance_df['DELAY'] + 
                                    performance_df['EXTENDED LOSS'] + 
                                    performance_df['PLANNED DOWN'] + 
                                    performance_df['STANDBY'] + 
                                    performance_df['UNPLANNED DOWN'])
    
    # Calculate PA using new formula: PA = (Operation hrs+DELAY+STANDBY) / (Grand Total-EXTENDED LOSS)*100
    performance_df['PA'] = ((performance_df['Operating hrs'] + performance_df['DELAY'] + performance_df['STANDBY']) / 
                          (performance_df['Grand Total'] - performance_df['EXTENDED LOSS'])) * 100
    
    # Calculate UA using formula: UA = (Operation hrs+DELAY) / (Operation hrs+DELAY+STANDBY)*100
    performance_df['UA'] = ((performance_df['Operating hrs'] + performance_df['DELAY']) / 
                          (performance_df['Operating hrs'] + performance_df['DELAY'] + performance_df['STANDBY'])) * 100
    
    # Round PA and UA to 1 decimal place
    performance_df['PA'] = performance_df['PA'].round(1)
    performance_df['UA'] = performance_df['UA'].round(1)
    
    # Add Grand Total row at the bottom with updated columns
    grand_total_row = pd.DataFrame({
        'Unit': ['Grand Total'],
        'Operating hrs': [performance_df['Operating hrs'].sum()],
        'DELAY': [performance_df['DELAY'].sum()],
        'EXTENDED LOSS': [performance_df['EXTENDED LOSS'].sum()],
        'PLANNED DOWN': [performance_df['PLANNED DOWN'].sum()],
        'STANDBY': [performance_df['STANDBY'].sum()],
        'UNPLANNED DOWN': [performance_df['UNPLANNED DOWN'].sum()],
        'Grand Total': [performance_df['Grand Total'].mean()],
        'PA': [performance_df['PA'].mean()],
        'UA': [performance_df['UA'].mean()]
    })
    
    # Append the Grand Total row
    performance_df = pd.concat([performance_df, grand_total_row], ignore_index=True)
    
    return performance_df

def process_excel_file(file):
    """
    Process Excel file with multiple sheets
    """
    global processing_status
    
    # Reset processing status
    processing_status = {
        "is_processing": True,
        "start_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
        "complete_time": None,
        "is_complete": False,
        "filename": file.filename,
        "sheets_processed": [],
        "error": None
    }
    
    # Dictionary to store sheet data
    result_dict = {}
    
    try:
        log_message(f"Memulai pemrosesan file: {file.filename}")
        
        # Check if the file has 'delay' and 'cycle' sheets
        xls = pd.ExcelFile(file)
        sheet_names = xls.sheet_names
        
        log_message(f"Sheets dalam file: {sheet_names}")
        
        delay_df = None
        cycle_df = None
        
        # Process all sheets
        for sheet_name in sheet_names:
            # Read the sheet data
            log_message(f"Membaca sheet: {sheet_name}")
            df = pd.read_excel(file, sheet_name=sheet_name)
            
            # Normalize Category column for better compatibility
            if 'Category' in df.columns:
                log_message(f"Standarisasi kolom Category pada sheet {sheet_name}")
                # Normalize category values
                df['Category'] = df['Category'].apply(lambda x: 
                    'DELAY' if pd.notna(x) and str(x).strip().upper() == 'DELAY'
                    else 'EXTENDED LOSS' if pd.notna(x) and str(x).strip().upper() in ['EXTENDED LOSS', 'EXTENDED']
                    else 'PLANNED DOWN' if pd.notna(x) and str(x).strip().upper() == 'PLANNED DOWN'
                    else 'STANDBY' if pd.notna(x) and str(x).strip().upper() == 'STANDBY'
                    else 'UNPLANNED DOWN' if pd.notna(x) and str(x).strip().upper() == 'UNPLANNED DOWN'
                    else x)
                
            log_message(f"Kolom pada sheet {sheet_name}: {df.columns.tolist()}")
            
            # Determine data type based on sheet name or column structure
            if sheet_name.lower() == "delay" or ('Start Date' in df.columns and 'Finish Date' in df.columns and 
                ('Machine' in df.columns or 'Unit' in df.columns) and 'Duration' in df.columns):
                data_type = 'delay'
                log_message(f"Sheet {sheet_name} terdeteksi sebagai delay")
                log_message(f"Membersihkan data delay pada sheet {sheet_name}")
                cleaned_df = clean_excel_data(df, data_type)
                result_dict[sheet_name] = cleaned_df
                delay_df = cleaned_df
                processing_status["sheets_processed"].append(f"{sheet_name} (delay)")
            elif sheet_name.lower() == "cycle" or (
                'Unit' in df.columns and 
                ('Start time' in df.columns or 'Start' in df.columns) and 
                ('Finish Time' in df.columns or 'Finish' in df.columns) and
                'Dur' in df.columns):
                data_type = 'cycle'
                log_message(f"Sheet {sheet_name} terdeteksi sebagai cycle")
                log_message(f"Membersihkan data cycle pada sheet {sheet_name}")
                cleaned_df = clean_excel_data(df, data_type)
                result_dict[sheet_name] = cleaned_df
                cycle_df = cleaned_df
                processing_status["sheets_processed"].append(f"{sheet_name} (cycle)")
            else:
                # Skip sheets that don't match expected formats
                log_message(f"Sheet {sheet_name} tidak terdeteksi sebagai format yang valid")
                continue
        
        # Create Performance sheet if we have both Delay and Cycle data
        if delay_df is not None and cycle_df is not None:
            log_message("Membuat sheet Performance berdasarkan data Delay dan Cycle")
            performance_df = create_performance_sheet(delay_df, cycle_df)
            if performance_df is not None:
                result_dict['Performance'] = performance_df
                processing_status["sheets_processed"].append("Performance")
    
    except Exception as e:
        error_msg = f"Error saat memproses file: {str(e)}"
        log_message(error_msg)
        processing_status["error"] = error_msg
        processing_status["is_processing"] = False
        raise Exception(f"Error processing file: {str(e)}")
    
    log_message(f"Hasil sheets yang diproses: {list(result_dict.keys())}")
    return result_dict

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'file' not in request.files:
            return render_template('index.html', error='No file part')
        
        file = request.files['file']
        
        if file.filename == '':
            return render_template('index.html', error='No selected file')
        
        if file and file.filename.endswith(('.xlsx', '.xls')):
            try:
                log_message(f"Menerima file upload: {file.filename}")
                
                # Process the Excel file with multiple sheets
                cleaned_data = process_excel_file(file)
                
                if not cleaned_data:
                    processing_status["error"] = "No valid data sheets found in the file"
                    processing_status["is_processing"] = False
                    return render_template('index.html', error='No valid data sheets found in the file')
                
                log_message("Membuat file Excel baru dengan data yang sudah dibersihkan")
                
                # Create a new Excel file with the cleaned data
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl', datetime_format='mm/dd/yyyy hh:mm:ss') as writer:
                    for sheet_name, df in cleaned_data.items():
                        log_message(f"Menulis sheet {sheet_name} ke file output")
                        
                        # Convert string dates to datetime objects for Excel compatibility
                        if 'Start' in df.columns:
                            try:
                                df['Start'] = pd.to_datetime(df['Start'])
                            except:
                                pass
                                
                        if 'Finish' in df.columns:
                            try:
                                df['Finish'] = pd.to_datetime(df['Finish'])
                            except:
                                pass
                        
                        # Write dataframe to Excel
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        
                        # Apply table formatting with colors based on sheet type
                        workbook = writer.book
                        worksheet = writer.sheets[sheet_name]
                        
                        # Get the dimensions of the data
                        max_row = len(df) + 1  # +1 for header
                        max_col = len(df.columns)
                        
                        # Create a table with filters using proper table styling
                        table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
                        
                        # Generate unique table name (Excel has restrictions on table names)
                        safe_sheet_name = ''.join(c for c in sheet_name if c.isalnum())
                        table_name = f"Table{safe_sheet_name}"
                        
                        # Select appropriate table style based on sheet type
                        table_style = "TableStyleMedium2"  # Default blue
                        if sheet_name.lower() == 'delay':
                            table_style = "TableStyleMedium3"  # Orange
                        elif sheet_name.lower() == 'cycle': 
                            table_style = "TableStyleMedium11"  # Green
                        elif sheet_name.lower() == 'performance':
                            table_style = "TableStyleMedium2"  # Blue
                            
                        # Create table with style and enable filters
                        tab = Table(displayName=table_name, ref=table_ref)
                        style = TableStyleInfo(name=table_style, showFirstColumn=False,
                                              showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                        tab.tableStyleInfo = style
                        
                        # Remove any existing tables to avoid conflicts
                        for tbl in worksheet._tables:
                            if tbl.displayName == table_name:
                                worksheet._tables.remove(tbl)
                                
                        worksheet.add_table(tab)
                        
                        # Add number formatting
                        if sheet_name.lower() in ['delay', 'cycle']:
                            # Format duration column to 2 decimals
                            if 'Dur' in df.columns:
                                dur_col_idx = df.columns.get_loc('Dur') + 1
                                dur_col_letter = get_column_letter(dur_col_idx)
                                
                                for row in range(2, max_row + 1):
                                    cell = worksheet[f"{dur_col_letter}{row}"]
                                    cell.number_format = '0.00'
                                    
                        elif sheet_name.lower() == 'performance':
                            # Format all numeric columns to 2 decimals
                            numeric_cols = ['Operating hrs', 'DELAY', 'EXTENDED LOSS', 'PLANNED DOWN', 'STANDBY', 'UNPLANNED DOWN']
                            for col_name in numeric_cols:
                                if col_name in df.columns:
                                    col_idx = df.columns.get_loc(col_name) + 1
                                    col_letter = get_column_letter(col_idx)
                                    
                                    for row in range(2, max_row + 1):
                                        cell = worksheet[f"{col_letter}{row}"]
                                        cell.number_format = '0.00'
                            
                            # Apply color formatting to category headers to match filter colors
                            if len(df.columns) > 0:
                                # Get header cells
                                for col_name in ['DELAY', 'EXTENDED LOSS', 'PLANNED DOWN', 'STANDBY', 'UNPLANNED DOWN']:
                                    if col_name in df.columns:
                                        col_idx = df.columns.get_loc(col_name) + 1
                                        header_cell = worksheet[f"{get_column_letter(col_idx)}1"]
                                        
                                        # Apply bold formatting
                                        header_cell.font = openpyxl.styles.Font(bold=True)
                            
                            # Format Grand Total column to 2 decimals
                            if 'Grand Total' in df.columns:
                                gt_col_idx = df.columns.get_loc('Grand Total') + 1
                                gt_col_letter = get_column_letter(gt_col_idx)
                                
                                for row in range(2, max_row + 1):
                                    cell = worksheet[f"{gt_col_letter}{row}"]
                                    cell.number_format = '0.00'
                            
                            # Format percentage columns (PA and UA)
                            for col_name in ['PA', 'UA']:
                                if col_name in df.columns:
                                    col_idx = df.columns.get_loc(col_name) + 1
                                    col_letter = get_column_letter(col_idx)
                                    
                                    for row in range(2, max_row + 1):
                                        cell = worksheet[f"{col_letter}{row}"]
                                        cell.number_format = '0.0"%"'
                            
                            # Make the Grand Total row stand out
                            if len(df) > 0:
                                for col in range(1, max_col + 1):
                                    cell = worksheet[max_row][col-1]
                                    cell.font = openpyxl.styles.Font(bold=True)
                        
                        # Auto-adjust column widths
                        for col in worksheet.columns:
                            max_length = 0
                            column = col[0].column_letter
                            for cell in col:
                                if cell.value:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                            adjusted_width = (max_length + 2)
                            worksheet.column_dimensions[column].width = min(adjusted_width, 30)  # Cap width at 30
                
                output.seek(0)
                
                # Update processing status
                processing_status["is_processing"] = False
                processing_status["is_complete"] = True
                processing_status["complete_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                log_message(f"Pemrosesan file selesai: {file.filename}")
                
                # Return the processed file
                return send_file(
                    output,
                    as_attachment=True,
                    download_name='cleaned_data.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            
            except Exception as e:
                error_msg = f"Error processing file: {str(e)}"
                log_message(error_msg)
                processing_status["error"] = error_msg
                processing_status["is_processing"] = False
                return render_template('index.html', error=error_msg)
        
        else:
            error_msg = "File must be an Excel file (.xlsx or .xls)"
            log_message(error_msg)
            return render_template('index.html', error=error_msg)
    
    return render_template('index.html')

@app.route('/processing-status')
def get_processing_status():
    """
    Endpoint to check the status of file processing
    """
    return jsonify(processing_status)

@app.route('/download-template')
def download_template():
    """
    Endpoint untuk mengunduh file template
    """
    template_path = os.path.join(os.getcwd(), 'Default Format Cycle & Delay.xlsx')
    
    try:
        return send_file(
            template_path,
            as_attachment=True,
            download_name='Template Cycle & Delay.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return render_template('index.html', error=f'Error downloading template: {str(e)}')

if __name__ == '__main__':
    # Ensure directories exist
    os.makedirs('static', exist_ok=True)
    os.makedirs('templates', exist_ok=True)
    app.run(debug=True, host='0.0.0.0', port=5050) 